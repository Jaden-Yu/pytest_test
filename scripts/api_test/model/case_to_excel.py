'''
    加载excel用例
'''
import sys
import openpyxl


def open_case_excel():
    case_list = []

    wb = openpyxl.load_workbook('../datas/demo.xlsx')  # 打开excel
    sheet = wb['demo']
    print(sheet.max_row)  # 最大行数
    print(sheet.max_column)  # 最大列数

    # 获取参数标题
    case_key = []
    for i in range(1, sheet.max_column + 1):
        case_key.append(sheet.cell(1, i).value)

    # 获取将excel中的用例添加到列表字典中
    row = 2
    while row <= sheet.max_row:
        case_row = {}
        for i in range(1, sheet.max_column + 1):
            case_row[case_key[i - 1]] = sheet.cell(row, i).value
        case_list.append(case_row)
        row += 1

    # 构造pytest参数化的集合
    test_case_list = []
    for i in range(0, len(case_list)):
        case = (
            case_list[i]['method'],
            case_list[i]['api'],
            (case_list[i]['request_params']),
            (case_list[i]['request_data']),
            (case_list[i]['headers'])
        )
        test_case_list.append(case)
    return test_case_list


if __name__ == "__main__":
    file_name = str(sys.argv[1])
    open_case_excel(file_name)
